"""Tests for the manual validation export/import pipeline."""
from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from low.validation.export import EXPORT_FILENAME, export_annotation_bundle
from low.validation.import_report import generate_validation_report
from low.validation.record_ids import country_speaker_id
from low.validation.sampler import load_raw_db, load_raw_sources, sample_all_tables


@pytest.fixture
def validation_db_path(minimal_db_path, tmp_path) -> Path:
    """Extend minimal DB with language names and extra speaker disagreement."""
    raw = json.loads(minimal_db_path.read_text(encoding="utf-8"))
    for row in raw["country_language_speakers"]:
        if (
            row["country_code"] == "RW"
            and row["language_code"] == "kin"
            and row["source"] == "cia"
        ):
            row["speaker_count"] = 2_000_000
            row["speaker_fraction"] = 0.145
    raw["language_names"] = [
        {
            "language_part3": "deu",
            "name": "Deutsch",
            "in_language_bcp47": "de",
            "in_language_part3": "deu",
            "script": "latn",
            "source": "LinguaMeta",
        },
        {
            "language_part3": "fra",
            "name": "French",
            "in_language_bcp47": "en",
            "in_language_part3": "eng",
            "script": "latn",
            "source": "LinguaMeta",
        },
    ]
    path = tmp_path / "low_db.json"
    path.write_text(json.dumps(raw), encoding="utf-8")

    sources_dir = tmp_path / "sources"
    sources_dir.mkdir()
    (sources_dir / "cia_speakers.json").write_text(
        json.dumps(
            [
                {
                    "country_code": "RW",
                    "country_name": "Rwanda",
                    "language_name": "Kinyarwanda",
                    "iso639_3": "kin",
                    "percent": 71.8,
                    "country_population": 13776698,
                    "speaker_count": 2000000,
                }
            ]
        ),
        encoding="utf-8",
    )
    (sources_dir / "cldr_speakers.json").write_text("[]", encoding="utf-8")
    (sources_dir / "linguameta_speakers.json").write_text("[]", encoding="utf-8")
    (sources_dir / "linguameta_names.json").write_text("[]", encoding="utf-8")
    (sources_dir / "linguameta_scripts.json").write_text("[]", encoding="utf-8")
    (sources_dir / "wikidata_speakers.json").write_text("[]", encoding="utf-8")
    (sources_dir / "low_scraper_speakers.json").write_text(
        json.dumps(
            [
                {
                    "country_code": "UG",
                    "iso639_3": "kin",
                    "speaker_count": 450000,
                    "speaker_fraction": 0.009838,
                }
            ]
        ),
        encoding="utf-8",
    )
    return path


def test_record_id_stability():
    row = {
        "country_code": "DE",
        "language_code": "deu",
        "source": "cldr",
        "speaker_count": 1,
        "speaker_fraction": 0.5,
    }
    assert country_speaker_id(row) == "speakers:DE:deu:cldr"
    assert country_speaker_id(row) == country_speaker_id(row)


def test_sample_includes_scraped_and_disagreement(validation_db_path):
    from low import LanguagesOfTheWorld

    raw = load_raw_db(validation_db_path)
    raw_sources = load_raw_sources(validation_db_path)
    db = LanguagesOfTheWorld(db_path=validation_db_path)

    tables, manifest = sample_all_tables(db, raw, raw_sources, seed=42)

    speaker_ids = {
        country_speaker_id(r) for r in tables["country_speakers"]
    }
    assert "speakers:UG:kin:scraped" in speaker_ids

    kin_rw_reasons = manifest["country_speakers"].get("speakers:RW:kin:cldr", [])
    kin_cia_reasons = manifest["country_speakers"].get("speakers:RW:kin:cia", [])
    assert any("disagreement" in r for r in kin_rw_reasons + kin_cia_reasons)

    lang_ids = {r["part3"] for r in tables["languages"]}
    assert "arc" in lang_ids  # missing glottocode risk


def test_export_reproducible_record_ids(validation_db_path, tmp_path):
    out1 = tmp_path / "run1"
    out2 = tmp_path / "run2"
    export_annotation_bundle(out1, db_path=validation_db_path, seed=42)
    export_annotation_bundle(out2, db_path=validation_db_path, seed=42)

    m1 = json.loads((out1 / "exports" / "manifest.json").read_text(encoding="utf-8"))
    m2 = json.loads((out2 / "exports" / "manifest.json").read_text(encoding="utf-8"))
    assert m1["tables"]["languages"]["records"] == m2["tables"]["languages"]["records"]


def test_export_writes_xlsx_workbook(validation_db_path, tmp_path):
    root = tmp_path / "annotation"
    exports_dir = export_annotation_bundle(root, db_path=validation_db_path, seed=42)

    xlsx_path = exports_dir / EXPORT_FILENAME
    assert xlsx_path.exists()
    assert (exports_dir / "manifest.json").exists()
    assert (root / "ANNOTATOR_GUIDE.md").exists()

    manifest = json.loads((exports_dir / "manifest.json").read_text(encoding="utf-8"))
    assert manifest["format"] == "xlsx"
    assert manifest["filename"] == EXPORT_FILENAME

    wb = load_workbook(xlsx_path)
    assert "countries" in wb.sheetnames
    assert "country_speakers" in wb.sheetnames
    assert "issue_categories" in wb.sheetnames
    assert "_options" in wb.sheetnames
    assert wb["_options"].sheet_state == "hidden"

    ws = wb["country_speakers"]
    headers = [cell.value for cell in ws[1]]
    assert "country_label" in headers
    assert "other_sources_json" in headers
    assert "valid" in headers
    assert "issue_category" in headers
    assert "annotator" not in headers
    assert "annotated_at" not in headers
    assert "severity" not in headers

    valid_col = headers.index("valid") + 1
    cell = ws.cell(row=2, column=valid_col)
    assert cell.fill.start_color.rgb in ("00D9D9D9", "FFD9D9D9")
    wb.close()


def test_export_data_validation_formulas_are_valid(validation_db_path, tmp_path):
    import re
    from zipfile import ZipFile

    root = tmp_path / "annotation"
    exports_dir = export_annotation_bundle(root, db_path=validation_db_path, seed=42)
    xlsx_path = exports_dir / EXPORT_FILENAME

    with ZipFile(xlsx_path) as z:
        for name in z.namelist():
            if not name.startswith("xl/worksheets/sheet") or not name.endswith(".xml"):
                continue
            xml = z.read(name).decode("utf-8")
            for formula in re.findall(r"<formula1>([^<]*)</formula1>", xml):
                assert '","' not in formula, f"malformed list formula in {name}: {formula!r}"
                assert not formula.startswith('="'), f"malformed list formula in {name}: {formula!r}"


def test_report_round_trip_xlsx(validation_db_path, tmp_path):
    root = tmp_path / "annotation"
    exports_dir = export_annotation_bundle(root, db_path=validation_db_path, seed=42)

    completed = root / "completed"
    completed.mkdir()
    src = exports_dir / EXPORT_FILENAME
    dst = completed / EXPORT_FILENAME
    dst.write_bytes(src.read_bytes())

    wb = load_workbook(dst)
    ws = wb["countries"]
    headers = [cell.value for cell in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers)}
    ws.cell(row=2, column=col["valid"]).value = "no"
    ws.cell(row=2, column=col["issue_category"]).value = "wrong_label"
    ws.cell(row=2, column=col["notes"]).value = "test flag"
    wb.save(dst)
    wb.close()

    report_dir = generate_validation_report(
        completed,
        root / "reports",
        manifest_path=exports_dir / "manifest.json",
    )

    summary = json.loads((report_dir / "summary.json").read_text(encoding="utf-8"))
    assert summary["totals"]["fail_count"] >= 1
    assert (report_dir / "summary.md").exists()
    assert (report_dir / "flagged_records.csv").exists()

    with (report_dir / "flagged_records.csv").open(encoding="utf-8") as fh:
        flagged = list(csv.DictReader(fh))
    assert any(r["notes"] == "test flag" for r in flagged)


def test_max_samples_caps_each_sheet(validation_db_path, tmp_path):
    from low import LanguagesOfTheWorld

    raw = load_raw_db(validation_db_path)
    raw_sources = load_raw_sources(validation_db_path)
    db = LanguagesOfTheWorld(db_path=validation_db_path)

    tables, _ = sample_all_tables(
        db, raw, raw_sources, seed=42, max_samples=3
    )
    for table, rows in tables.items():
        assert len(rows) <= 3, f"{table} has {len(rows)} rows"


def test_export_max_samples_in_manifest(validation_db_path, tmp_path):
    root = tmp_path / "annotation"
    export_annotation_bundle(
        root, db_path=validation_db_path, seed=42, max_samples=5
    )
    manifest = json.loads(
        (root / "exports" / "manifest.json").read_text(encoding="utf-8")
    )
    assert manifest["max_samples"] == 5
    for table_info in manifest["tables"].values():
        assert table_info["row_count"] <= 5
