"""XLSX export with context enrichment and color-coded annotation dropdowns."""
from __future__ import annotations

import json
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from low import LanguagesOfTheWorld

from .record_ids import ID_FN_BY_TABLE, add_record_id, other_sources_json
from .sampler import (
    ANNOTATION_COLUMNS,
    load_raw_db,
    load_raw_sources,
    sample_all_tables,
)

EXPORT_FILENAME = "low_validation.xlsx"
INSTRUCTIONS_SHEET = "Instructions"
LEGEND_SHEET = "issue_categories"
OPTIONS_SHEET = "_options"
ANNOTATION_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")

VALID_VALUES = ["yes", "no", "unsure"]

VALID_COLORS = {
    "yes": "FFC6EFCE",
    "no": "FFFFC7CE",
    "unsure": "FFFFEB9C",
}

# Color palette for issue categories, grouped by error type.
ISSUE_CATEGORY_COLORS: dict[str, str] = {
    "wrong_count": "FFE8D4F0",
    "wrong_country": "FFD4B8E8",
    "wrong_language": "FFC49BDB",
    "fraction_mismatch": "FFB07CC6",
    "source_error": "FF9B59B6",
    "wrong_label": "FFBDD7EE",
    "wrong_code": "FF9DC3E6",
    "missing_glottocode": "FF7FB3D5",
    "implausible_speakers": "FF5B9BD5",
    "wrong_script": "FFC5E0B4",
    "not_a_name": "FFA9D18E",
    "canonical_error": "FF8DBE6B",
    "wrong_parent": "FFF8CBAD",
    "orphan_node": "FFF4B183",
    "wrong_hierarchy": "FFED7D31",
    "implausible_population": "FFE36C09",
    "wrong_status": "FFFFD966",
    "wrong_canonical": "FFFFC000",
    "merge_mismatch": "FFF4CCCC",
    "wrong_mapping": "FFE6B8B8",
    "upstream_error": "FFD99694",
}

DROPDOWN_ROWS = [
    ("*", "valid", "yes"),
    ("*", "valid", "no"),
    ("*", "valid", "unsure"),
    ("country_speakers", "issue_category", "wrong_count"),
    ("country_speakers", "issue_category", "wrong_country"),
    ("country_speakers", "issue_category", "wrong_language"),
    ("country_speakers", "issue_category", "fraction_mismatch"),
    ("country_speakers", "issue_category", "source_error"),
    ("languages", "issue_category", "wrong_label"),
    ("languages", "issue_category", "wrong_code"),
    ("languages", "issue_category", "missing_glottocode"),
    ("languages", "issue_category", "implausible_speakers"),
    ("language_names", "issue_category", "wrong_script"),
    ("language_names", "issue_category", "not_a_name"),
    ("language_names", "issue_category", "wrong_language"),
    ("language_names", "issue_category", "canonical_error"),
    ("families", "issue_category", "wrong_parent"),
    ("families", "issue_category", "wrong_label"),
    ("families", "issue_category", "orphan_node"),
    ("countries", "issue_category", "wrong_label"),
    ("countries", "issue_category", "wrong_hierarchy"),
    ("countries", "issue_category", "implausible_population"),
    ("regions", "issue_category", "wrong_label"),
    ("regions", "issue_category", "wrong_hierarchy"),
    ("continents", "issue_category", "wrong_label"),
    ("continents", "issue_category", "wrong_hierarchy"),
    ("official_languages", "issue_category", "wrong_status"),
    ("official_languages", "issue_category", "wrong_country"),
    ("official_languages", "issue_category", "wrong_language"),
    ("language_scripts", "issue_category", "wrong_script"),
    ("language_scripts", "issue_category", "wrong_canonical"),
    ("language_scripts", "issue_category", "wrong_label"),
    ("scripts", "issue_category", "wrong_label"),
    ("scripts", "issue_category", "wrong_code"),
    ("raw_*", "issue_category", "merge_mismatch"),
    ("raw_*", "issue_category", "wrong_mapping"),
    ("raw_*", "issue_category", "upstream_error"),
]

SHEET_NAMES = {
    "continents": "continents",
    "regions": "regions",
    "countries": "countries",
    "families": "families",
    "languages": "languages",
    "scripts": "scripts",
    "language_scripts": "language_scripts",
    "country_speakers": "country_speakers",
    "official_languages": "official_languages",
    "language_names": "language_names",
    "raw_cldr_speakers": "raw_cldr_speakers",
    "raw_cia_speakers": "raw_cia_speakers",
    "raw_linguameta_speakers": "raw_linguameta_speakers",
    "raw_linguameta_names": "raw_linguameta_names",
    "raw_linguameta_scripts": "raw_linguameta_scripts",
    "raw_wikidata_speakers": "raw_wikidata_speakers",
    "raw_scraper_speakers": "raw_scraper_speakers",
}


def _dataset_key(table: str) -> str:
    if table.startswith("raw_"):
        return "raw_*"
    return table


def _issue_categories_for_table(table: str) -> list[str]:
    key = _dataset_key(table)
    return [
        value
        for dataset, field, value in DROPDOWN_ROWS
        if field == "issue_category" and dataset in (key, "*")
    ]


def _family_path(db: LanguagesOfTheWorld, glottocode: str | None) -> str:
    if not glottocode:
        return ""
    fam = db.families.get(glottocode)
    if fam is None:
        return ""
    parts = [fam.label]
    node = fam.parent
    while node is not None:
        parts.append(node.label)
        node = node.parent
    return " > ".join(reversed(parts))


def _enrich_row(
    table: str,
    row: dict[str, Any],
    db: LanguagesOfTheWorld,
    speaker_index: dict[tuple[str, str], list[dict[str, Any]]],
    scripts_by_code: dict[str, str],
) -> dict[str, Any]:
    enriched = dict(row)

    if table == "continents":
        enriched["country_count"] = len(
            [c for c in db.countries if c.continent.id == row["id"]]
        )

    elif table == "regions":
        reg = db.regions.get(row["id"])
        enriched["continent_label"] = reg.continent.label if reg else ""
        enriched["country_count"] = len(reg.countries) if reg else 0

    elif table == "countries":
        country = db.countries.get(row["code"])
        if country:
            enriched["continent_label"] = country.continent.label
            enriched["region_label"] = country.region.label
            enriched["language_count"] = len(country.languages)

    elif table == "families":
        fam = db.families.get(row["glottocode"])
        if fam:
            enriched["parent_label"] = fam.parent.label if fam.parent else ""
            enriched["depth"] = fam.depth
            enriched["child_count"] = len(fam.children)
            enriched["language_count"] = len(fam.languages)
            enriched["family_path"] = _family_path(db, row["glottocode"])

    elif table == "languages":
        lang = db.languages.get(row["part3"])
        if lang:
            enriched["family_path"] = _family_path(db, lang.glottocode)
            enriched["country_count"] = len(lang.countries)
            enriched["has_speaker_data"] = bool(lang.speaker_counts)
            enriched["country_codes_display"] = ",".join(c.code for c in lang.countries)

    elif table == "language_scripts":
        lang = db.languages.get(row["language_part3"])
        script_label = scripts_by_code.get(row["script_code"], "")
        enriched["language_label"] = lang.label if lang else ""
        enriched["script_label"] = script_label

    elif table == "country_speakers":
        country = db.countries.get(row["country_code"])
        lang = db.languages.get(row["language_code"])
        key = (row["country_code"], row["language_code"])
        others = [
            {
                "source": s["source"],
                "speaker_count": s.get("speaker_count"),
                "speaker_fraction": s.get("speaker_fraction"),
            }
            for s in speaker_index.get(key, [])
            if s["source"] != row["source"]
        ]
        frac = float(row.get("speaker_fraction") or 0.0)
        enriched["country_label"] = country.label if country else ""
        enriched["language_label"] = lang.label if lang else ""
        enriched["country_population"] = country.population if country else 0
        enriched["fraction_pct"] = round(frac * 100, 4)
        enriched["other_sources_json"] = other_sources_json(others)

    elif table == "official_languages":
        country = db.countries.get(row["country_code"])
        lang = db.languages.get(row["language_code"])
        enriched["country_label"] = country.label if country else ""
        enriched["language_label"] = lang.label if lang else ""
        if country:
            enriched["continent_label"] = country.continent.label

    elif table == "language_names":
        lang = db.languages.get(row["language_part3"])
        enriched["language_label"] = lang.label if lang else ""
        script_code = row.get("script") or ""
        enriched["name_script_label"] = scripts_by_code.get(script_code, script_code)

    return enriched


def _build_speaker_index(
    speaker_rows: list[dict[str, Any]],
) -> dict[tuple[str, str], list[dict[str, Any]]]:
    index: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in speaker_rows:
        key = (row["country_code"], row["language_code"])
        index.setdefault(key, []).append(row)
    return index


def _row_to_export(
    table: str,
    row: dict[str, Any],
    sample_reasons: list[str],
) -> dict[str, Any]:
    out = add_record_id(table, row)
    out["sample_reason"] = ";".join(sample_reasons)
    for col in ANNOTATION_COLUMNS:
        out.setdefault(col, "")
    return out


def _fieldnames(rows: list[dict[str, Any]]) -> list[str]:
    names: list[str] = []
    for row in rows:
        for key in row:
            if key not in names:
                names.append(key)
    return names


def _cell_export_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    return value


def _write_options_sheet(ws: Worksheet) -> dict[tuple[str, str], tuple[int, int]]:
    """Write hidden lookup ranges for per-dataset dropdown values."""
    ws.append(["dataset", "field", "value", "color"])
    range_map: dict[tuple[str, str], tuple[int, int]] = {}
    current_dataset = ""
    current_field = ""
    block_start = 2

    for row_idx, (dataset, field, value) in enumerate(DROPDOWN_ROWS, start=2):
        color = ""
        if field == "issue_category":
            color = ISSUE_CATEGORY_COLORS.get(value, "FFE0E0E0")
        elif field == "valid":
            color = VALID_COLORS.get(value, "")

        ws.append([dataset, field, value, color])
        if color:
            ws.cell(row=row_idx, column=4).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

        key = (dataset, field)
        if key != (current_dataset, current_field):
            if current_dataset:
                range_map[(current_dataset, current_field)] = (block_start, row_idx - 1)
            current_dataset, current_field = dataset, field
            block_start = row_idx

    if current_dataset:
        range_map[(current_dataset, current_field)] = (block_start, row_idx)

    ws.sheet_state = "hidden"
    return range_map


def _write_legend_sheet(ws: Worksheet) -> None:
    ws.append(["issue_category", "color", "description"])
    descriptions = {
        "wrong_count": "Speaker count is incorrect",
        "wrong_country": "Country assignment is wrong",
        "wrong_language": "Language assignment is wrong",
        "fraction_mismatch": "Fraction does not match count ÷ population",
        "source_error": "Source attribution is wrong",
        "wrong_label": "Display label is wrong",
        "wrong_code": "ISO or other code is wrong",
        "missing_glottocode": "Expected Glottolog code is absent",
        "implausible_speakers": "Speaker count is implausible",
        "wrong_script": "Script code or assignment is wrong",
        "not_a_name": "Value is not a valid language name",
        "canonical_error": "Canonical name assignment is wrong",
        "wrong_parent": "Family parent link is wrong",
        "orphan_node": "Family node has no valid parent",
        "wrong_hierarchy": "Geographic hierarchy is wrong",
        "implausible_population": "Population figure is implausible",
        "wrong_status": "Official-language status is wrong",
        "wrong_canonical": "Canonical script flag is wrong",
        "merge_mismatch": "LOW merge does not match upstream",
        "wrong_mapping": "Code mapping during merge is wrong",
        "upstream_error": "Error originates in upstream source",
    }
    for category in sorted(ISSUE_CATEGORY_COLORS):
        color = ISSUE_CATEGORY_COLORS[category]
        ws.append([category, "", descriptions.get(category, "")])
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=2).fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 48
    ws.freeze_panes = "A2"


def _write_instructions_sheet(ws: Worksheet) -> None:
    lines = [
        "LOW Manual Validation",
        "",
        "Open ANNOTATOR_GUIDE.md (in the annotation folder) for full instructions.",
        "",
        "Quick steps:",
        "1. Work through each data sheet (tabs at the bottom).",
        "2. For each row, read the context columns and fill in the light-gray cells.",
        "3. Use the dropdowns in valid and issue_category.",
        "4. Add notes with evidence when you mark a row as no or unsure.",
        "5. Save the workbook and return it to the project lead.",
        "",
        "See the issue_categories tab for a color key.",
    ]
    for line in lines:
        ws.append([line])
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 80


def _options_range(
    range_map: dict[tuple[str, str], tuple[int, int]],
    dataset: str,
    field: str,
) -> str | None:
    bounds = range_map.get((dataset, field))
    if bounds is None:
        return None
    start, end = bounds
    return f"'{OPTIONS_SHEET}'!$C${start}:$C${end}"


def _add_dropdown(
    ws: Worksheet,
    col_idx: int,
    last_row: int,
    formula: str,
    *,
    allow_blank: bool = True,
) -> None:
    if last_row < 2:
        return
    col_letter = get_column_letter(col_idx)
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=allow_blank,
        showDropDown=False,
    )
    dv.error = "Please choose a value from the dropdown list."
    dv.errorTitle = "Invalid value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def _add_value_coloring(
    ws: Worksheet,
    col_idx: int,
    last_row: int,
    value_colors: dict[str, str],
) -> None:
    if last_row < 2:
        return
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{last_row}"
    for value, color in value_colors.items():
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )


def _add_issue_category_coloring(
    ws: Worksheet,
    col_idx: int,
    last_row: int,
    categories: list[str],
) -> None:
    if last_row < 2:
        return
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{last_row}"
    for category in categories:
        color = ISSUE_CATEGORY_COLORS.get(category, "FFE0E0E0")
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[f'${col_letter}2="{category}"'],
                fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
            ),
        )


def _write_data_sheet(
    ws: Worksheet,
    table: str,
    rows: list[dict[str, Any]],
    range_map: dict[tuple[str, str], tuple[int, int]],
) -> None:
    if not rows:
        return

    headers = _fieldnames(rows)
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = ANNOTATION_FILL if header in ANNOTATION_COLUMNS else header_fill
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([_cell_export_value(row.get(h, "")) for h in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_index = {name: idx + 1 for idx, name in enumerate(headers)}
    last_row = len(rows) + 1
    dataset_key = _dataset_key(table)

    for header in ANNOTATION_COLUMNS:
        if header not in col_index:
            continue
        col = col_index[header]
        for row_idx in range(2, last_row + 1):
            ws.cell(row=row_idx, column=col).fill = ANNOTATION_FILL

    if "valid" in col_index:
        formula = _options_range(range_map, "*", "valid")
        if formula:
            _add_dropdown(ws, col_index["valid"], last_row, formula)
            _add_value_coloring(ws, col_index["valid"], last_row, VALID_COLORS)

    if "issue_category" in col_index:
        categories = _issue_categories_for_table(table)
        formula = _options_range(range_map, dataset_key, "issue_category")
        if formula and categories:
            _add_dropdown(ws, col_index["issue_category"], last_row, formula)
            _add_issue_category_coloring(
                ws, col_index["issue_category"], last_row, categories
            )

    for col_idx, header in enumerate(headers, start=1):
        width = min(40, max(10, len(header) + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_workbook(
    tables: dict[str, list[dict[str, Any]]],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    instructions = wb.create_sheet(INSTRUCTIONS_SHEET, 0)
    _write_instructions_sheet(instructions)

    legend = wb.create_sheet(LEGEND_SHEET, 1)
    _write_legend_sheet(legend)

    options = wb.create_sheet(OPTIONS_SHEET)
    range_map = _write_options_sheet(options)

    for table, rows in tables.items():
        sheet_name = SHEET_NAMES[table]
        ws = wb.create_sheet(sheet_name)
        _write_data_sheet(ws, table, rows, range_map)

    return wb


def export_annotation_bundle(
    output_dir: Path,
    *,
    db_path: Path | None = None,
    seed: int = 42,
    include_risk: bool = True,
    max_samples: int | None = None,
) -> Path:
    """Export stratified annotation workbook and supporting files."""
    if db_path is None:
        db_path = Path(__file__).resolve().parents[1] / "data" / "low_db.json"

    output_dir = Path(output_dir)
    exports_dir = output_dir / "exports"
    exports_dir.mkdir(parents=True, exist_ok=True)

    raw = load_raw_db(db_path)
    raw_sources = load_raw_sources(db_path)
    db = LanguagesOfTheWorld(db_path=db_path)

    tables, manifest = sample_all_tables(
        db, raw, raw_sources, seed=seed, include_risk=include_risk, max_samples=max_samples
    )

    speaker_index = _build_speaker_index(raw.get("country_language_speakers", []))
    scripts_by_code = {s.code: s.label for s in db.scripts}

    export_tables: dict[str, list[dict[str, Any]]] = {}
    manifest_out: dict[str, Any] = {
        "seed": seed,
        "include_risk": include_risk,
        "max_samples": max_samples,
        "db_path": str(db_path.resolve()),
        "format": "xlsx",
        "filename": EXPORT_FILENAME,
        "tables": {},
    }

    for table, rows in tables.items():
        table_manifest = manifest.get(table, {})
        export_rows = []
        for row in rows:
            reasons = table_manifest.get(ID_FN_BY_TABLE[table](row), ["stratified"])
            enriched = _enrich_row(table, row, db, speaker_index, scripts_by_code)
            export_rows.append(_row_to_export(table, enriched, reasons))

        export_tables[table] = export_rows
        manifest_out["tables"][table] = {
            "sheet": SHEET_NAMES[table],
            "row_count": len(export_rows),
            "records": {
                r["record_id"]: r["sample_reason"] for r in export_rows
            },
        }

    xlsx_path = exports_dir / EXPORT_FILENAME
    workbook = _write_workbook(export_tables)
    workbook.save(xlsx_path)

    manifest_path = exports_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest_out, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    guide_src = Path(__file__).resolve().parents[3] / "annotation" / "ANNOTATOR_GUIDE.md"
    guide_dst = output_dir / "ANNOTATOR_GUIDE.md"
    if guide_src.exists() and guide_src.resolve() != guide_dst.resolve():
        shutil.copy2(guide_src, guide_dst)

    return exports_dir
