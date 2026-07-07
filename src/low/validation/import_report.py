"""Parse completed annotation workbooks and generate validation reports."""
from __future__ import annotations

import csv
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SKIP_SHEETS = {"Instructions", "issue_categories", "_options"}
DEFAULT_WORKBOOK = "low_validation.xlsx"


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _read_annotation_xlsx(path: Path) -> dict[str, list[dict[str, str]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    datasets: dict[str, list[dict[str, str]]] = {}
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [_cell_value(h) for h in rows[0]]
        if not headers or not headers[0]:
            continue
        data_rows: list[dict[str, str]] = []
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            data_rows.append(
                {
                    headers[i]: _cell_value(row[i]) if i < len(row) else ""
                    for i in range(len(headers))
                }
            )
        datasets[ws.title] = data_rows
    wb.close()
    return datasets


def _read_annotation_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists() or path.stat().st_size == 0:
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return list(csv.DictReader(fh))


def _summarize_rows(rows: list[dict[str, str]], dataset: str) -> dict[str, Any]:
    total = len(rows)
    annotated = [r for r in rows if (r.get("valid") or "").strip()]
    valid_counts = Counter((r.get("valid") or "").strip().lower() for r in annotated)
    issue_counts = Counter(
        (r.get("issue_category") or "").strip()
        for r in annotated
        if (r.get("valid") or "").strip().lower() == "no"
    )
    yes = valid_counts.get("yes", 0)
    no = valid_counts.get("no", 0)
    unsure = valid_counts.get("unsure", 0)
    reviewed = yes + no + unsure
    pass_rate = (yes / reviewed) if reviewed else None

    return {
        "dataset": dataset,
        "total_rows": total,
        "annotated_rows": len(annotated),
        "reviewed_rows": reviewed,
        "pass_count": yes,
        "fail_count": no,
        "unsure_count": unsure,
        "pass_rate": pass_rate,
        "top_issue_categories": issue_counts.most_common(10),
    }


def _collect_flagged(rows: list[dict[str, str]], dataset: str) -> list[dict[str, str]]:
    flagged = []
    for row in rows:
        valid = (row.get("valid") or "").strip().lower()
        if valid == "no":
            flagged.append(
                {
                    "dataset": dataset,
                    "record_id": row.get("record_id", ""),
                    "issue_category": row.get("issue_category", ""),
                    "notes": row.get("notes", ""),
                    "sample_reason": row.get("sample_reason", ""),
                }
            )
    flagged.sort(key=lambda r: (r["dataset"], r["record_id"]))
    return flagged


def _load_completed_datasets(input_dir: Path) -> dict[str, list[dict[str, str]]]:
    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    if xlsx_files:
        return _read_annotation_xlsx(xlsx_files[0])

    datasets: dict[str, list[dict[str, str]]] = {}
    for path in sorted(input_dir.glob("*.csv")):
        if path.name == "dropdown_values.csv":
            continue
        datasets[path.stem] = _read_annotation_csv(path)
    return datasets


def generate_validation_report(
    input_dir: Path,
    output_dir: Path,
    *,
    manifest_path: Path | None = None,
) -> Path:
    """Aggregate completed annotation files into summary and flagged-record reports."""
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest: dict[str, Any] = {}
    if manifest_path and manifest_path.exists():
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))

    datasets = _load_completed_datasets(input_dir)
    per_dataset: list[dict[str, Any]] = []
    all_flagged: list[dict[str, str]] = []

    for dataset, rows in sorted(datasets.items()):
        per_dataset.append(_summarize_rows(rows, dataset))
        all_flagged.extend(_collect_flagged(rows, dataset))

    generated_at = datetime.now(timezone.utc).isoformat()
    summary = {
        "generated_at": generated_at,
        "input_dir": str(input_dir.resolve()),
        "manifest_path": str(manifest_path.resolve()) if manifest_path else None,
        "datasets": per_dataset,
        "totals": {
            "datasets": len(per_dataset),
            "total_rows": sum(d["total_rows"] for d in per_dataset),
            "annotated_rows": sum(d["annotated_rows"] for d in per_dataset),
            "fail_count": sum(d["fail_count"] for d in per_dataset),
            "unsure_count": sum(d["unsure_count"] for d in per_dataset),
            "flagged_records": len(all_flagged),
        },
    }

    if manifest:
        summary["export_seed"] = manifest.get("seed")
        summary["export_db_path"] = manifest.get("db_path")
        summary["export_format"] = manifest.get("format")

    json_path = output_dir / "summary.json"
    json_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )

    md_lines = [
        "# LOW Manual Validation Report",
        "",
        f"Generated: {generated_at}",
        "",
        "## Totals",
        "",
        f"- Datasets reviewed: {summary['totals']['datasets']}",
        f"- Total export rows: {summary['totals']['total_rows']}",
        f"- Annotated rows: {summary['totals']['annotated_rows']}",
        f"- Failed (`valid=no`): {summary['totals']['fail_count']}",
        f"- Unsure: {summary['totals']['unsure_count']}",
        f"- Flagged records: {summary['totals']['flagged_records']}",
        "",
        "## Per-dataset summary",
        "",
        "| Dataset | Rows | Reviewed | Pass | Fail | Unsure | Pass rate |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for d in per_dataset:
        rate = d["pass_rate"]
        rate_str = f"{rate * 100:.1f}%" if rate is not None else "—"
        md_lines.append(
            f"| {d['dataset']} | {d['total_rows']} | {d['reviewed_rows']} | "
            f"{d['pass_count']} | {d['fail_count']} | {d['unsure_count']} | {rate_str} |"
        )

    md_lines.extend(["", "## Top issue categories (failed rows)", ""])
    for d in per_dataset:
        if not d["top_issue_categories"]:
            continue
        md_lines.append(f"### {d['dataset']}")
        for category, count in d["top_issue_categories"]:
            label = category or "(unspecified)"
            md_lines.append(f"- {label}: {count}")
        md_lines.append("")

    md_path = output_dir / "summary.md"
    md_path.write_text("\n".join(md_lines) + "\n", encoding="utf-8")

    flagged_path = output_dir / "flagged_records.csv"
    if all_flagged:
        fieldnames = list(all_flagged[0].keys())
        with flagged_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_flagged)
    else:
        flagged_path.write_text("", encoding="utf-8")

    return output_dir
